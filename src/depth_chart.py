"""
depth_chart.py
--------------
Load 2026 depth chart data from per-team Excel files.

The depth chart Excel files live in a directory named "2026 Depth Chart"
and are named like "2025-Yankees-Depth-Charts.xlsx".

Public API
----------
get_depth_chart_dir(data_dir)                        -> str | None
load_projected_roster(depth_chart_dir, team)         -> pd.DataFrame
load_minors_players(depth_chart_dir, team, levels)   -> pd.DataFrame
"""

from __future__ import annotations

import os

import numpy as np
import pandas as pd

# -----------------------------------------------------------------------
# Team abbreviation → depth chart filename
# -----------------------------------------------------------------------
TEAM_FILE_MAP: dict[str, str] = {
    "ARI": "2025-Diamondbacks-Depth-Charts.xlsx",
    "ATH": "2025-Athletics-Depth-Charts.xlsx",
    "ATL": "2025-Braves-Depth-Charts.xlsx",
    "BAL": "2025-Orioles-Depth-Charts.xlsx",
    "BOS": "2025-Red Sox-Depth-Charts.xlsx",
    "CHC": "2025-Cubs-Depth-Charts.xlsx",
    "CHW": "2025-White Sox-Depth-Charts.xlsx",
    "CIN": "2025-Reds-Depth-Charts.xlsx",
    "CLE": "2025-Guardians-Depth-Charts.xlsx",
    "COL": "2025-Rockies-Depth-Charts.xlsx",
    "DET": "2025-Tigers-Depth-Charts.xlsx",
    "HOU": "2025-Astros-Depth-Charts.xlsx",
    "KCR": "2025-Royals-Depth-Charts.xlsx",
    "LAA": "2025-Angels-Depth-Charts.xlsx",
    "LAD": "2025-Dodgers-Depth-Charts.xlsx",
    "MIA": "2025-Marlins-Depth-Charts.xlsx",
    "MIL": "2025-Brewers-Depth-Charts.xlsx",
    "MIN": "2025-Twins-Depth-Charts.xlsx",
    "NYM": "2025-Mets-Depth-Charts.xlsx",
    "NYY": "2025-Yankees-Depth-Charts.xlsx",
    "PHI": "2025-Phillies-Depth-Charts.xlsx",
    "PIT": "2025-Pirates-Depth-Charts.xlsx",
    "SDP": "2025-Padres-Depth-Charts.xlsx",
    "SEA": "2025-Mariners-Depth-Charts.xlsx",
    "SFG": "2025-Giants-Depth-Charts.xlsx",
    "STL": "2025-Cardinals-Depth-Charts.xlsx",
    "TBR": "2025-Rays-Depth-Charts.xlsx",
    "TEX": "2025-Rangers-Depth-Charts.xlsx",
    "TOR": "2025-Blue Jays-Depth-Charts.xlsx",
    "WSN": "2025-Nationals-Depth-Charts.xlsx",
}

# The 4 "Projected" sheets that represent the active 26-man roster
_PROJECTED_SHEETS = [
    "Projected Go-To Starting Lineup",
    "Projected Bench",
    "Projected Starting Rotation",
    "Projected Bullpen",
]

# Position-level depth chart sheets (minor leagues)
_MINORS_SHEETS = ["C", "1B", "2B", "3B", "SS", "OF", "SP", "RP"]

# 2026 MLB league minimum salary in $M
LEAGUE_MIN_M: float = 0.74

# Position group mapping (mirrors team_mode.py / projections.py)
_POS_GROUP_MAP = {
    "C":   "C",
    "1B":  "1B",
    "3B":  "3B",
    "IF":  "1B",
    "2B":  "2B",
    "SS":  "SS",
    "LF":  "OF",
    "CF":  "OF",
    "RF":  "OF",
    "OF":  "OF",
    "SP":  "SP",
    "RP":  "RP",
    "TWP": "SP",
    "DH":  "DH",
    "P":   "SP",
}


def _map_pos(pos_raw: str) -> str:
    """Map raw POS string (possibly 'CF/LF') to canonical pos_group."""
    if not pos_raw or pd.isna(pos_raw):
        return "UNK"
    primary = str(pos_raw).strip().split("/")[0].upper()
    return _POS_GROUP_MAP.get(primary, "UNK")


def _is_40man_option(val) -> bool:
    """Return True if Options/R5 value is numeric (0–3) → player is on 40-man roster.

    'R5', "Dec'XX", 'n/a', None → not on 40-man (returns False).
    """
    if val is None:
        return False
    s = str(val).strip()
    try:
        n = int(float(s))
        return 0 <= n <= 3
    except (ValueError, TypeError):
        return False


def get_depth_chart_dir(data_dir: str) -> str | None:
    """
    Find the depth chart directory by searching several candidate locations.

    Search order:
      1. <data_dir>/2026 Depth Chart/
      2. Walk up the directory tree looking for a sibling mlb_optimizer2
    """
    candidates = [
        os.path.join(data_dir, "2026 Depth Chart"),
    ]

    # Walk up to 6 levels from data_dir and check for sibling mlb_optimizer2
    base = os.path.abspath(data_dir)
    for _ in range(6):
        parent = os.path.dirname(base)
        if parent == base:
            break  # reached filesystem root
        sibling = os.path.join(parent, "mlb_optimizer2", "mlb_optimizer", "data", "2026 Depth Chart")
        candidates.append(sibling)
        # Also try direct sibling (parent/2026 Depth Chart)
        candidates.append(os.path.join(parent, "data", "2026 Depth Chart"))
        base = parent

    for path in candidates:
        if os.path.isdir(path):
            # Verify at least one team file exists
            for fname in TEAM_FILE_MAP.values():
                if os.path.isfile(os.path.join(path, fname)):
                    return path
    return None


def _read_projected_sheet(
    wb,
    sheet_name: str,
) -> pd.DataFrame:
    """
    Read one projected sheet and return normalised rows.

    Returns DataFrame with columns:
        Player, pos_raw, pos_group, age, proj_WAR, depth_sheet
    """
    if sheet_name not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return pd.DataFrame()

    header = [str(c).strip() if c else "" for c in rows[0]]

    def _col(name: str) -> int | None:
        for i, h in enumerate(header):
            if h.upper() == name.upper():
                return i
        return None

    idx_player  = _col("PLAYER")
    idx_pos     = _col("POS")
    idx_age     = _col("AGE")
    idx_war     = _col("WAR")

    if idx_player is None:
        return pd.DataFrame()

    records = []
    for row in rows[1:]:
        if not row or row[idx_player] is None:
            continue
        player = str(row[idx_player]).strip()
        if not player or player.lower() in ("player", "nan", ""):
            continue

        pos_raw  = str(row[idx_pos]).strip() if idx_pos is not None and row[idx_pos] else "UNK"
        age      = float(row[idx_age]) if idx_age is not None and row[idx_age] not in (None, "") else np.nan
        proj_war = float(row[idx_war]) if idx_war is not None and row[idx_war] not in (None, "") else np.nan

        records.append({
            "Player":       player,
            "pos_raw":      pos_raw,
            "pos_group":    _map_pos(pos_raw),
            "age":          age,
            "proj_WAR":     proj_war,
            "depth_sheet":  sheet_name,
        })

    return pd.DataFrame(records)


def load_projected_roster(depth_chart_dir: str, team: str) -> pd.DataFrame:
    """
    Load all players from the 4 "Projected" tabs of the team's depth chart.

    Returns DataFrame with columns:
        Player, pos_raw, pos_group, age, proj_WAR, depth_sheet

    Players with unknown positions are included but flagged as pos_group='UNK'.
    Duplicate players (e.g. multi-position) are deduplicated keeping the
    row with the highest proj_WAR.
    """
    fname = TEAM_FILE_MAP.get(team.upper())
    if fname is None:
        return pd.DataFrame()

    fpath = os.path.join(depth_chart_dir, fname)
    if not os.path.isfile(fpath):
        return pd.DataFrame()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()

    parts = [_read_projected_sheet(wb, s) for s in _PROJECTED_SHEETS]
    df = pd.concat([p for p in parts if not p.empty], ignore_index=True)

    if df.empty:
        return df

    # Deduplicate: keep highest proj_WAR row per player
    df = (
        df.sort_values("proj_WAR", ascending=False, na_position="last", kind="mergesort")
          .drop_duplicates("Player", keep="first")
          .reset_index(drop=True)
    )
    return df


def load_minors_players(
    depth_chart_dir: str,
    team: str,
    levels: tuple[str, ...] = ("AAA", "AA"),
) -> pd.DataFrame:
    """
    Load minor-league / depth players from the position-level sheets.

    Only includes players whose PROJ LEVEL is in ``levels``.

    Returns DataFrame with columns:
        Player, pos_raw, pos_group, age, proj_level, max_level, on_40_man, depth_sheet

    ``on_40_man`` is True when the 'Options or R5 Status' column contains a
    numeric value 0-3 (player is on the 40-man roster, optioned to minors).
    R5 dates, 'n/a', and blank values indicate the player is NOT yet on the 40-man.
    """
    fname = TEAM_FILE_MAP.get(team.upper())
    if fname is None:
        return pd.DataFrame()

    fpath = os.path.join(depth_chart_dir, fname)
    if not os.path.isfile(fpath):
        return pd.DataFrame()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()

    records = []
    for sheet_name in _MINORS_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        header = [str(c).strip() if c else "" for c in rows[0]]

        def _col(name: str) -> int | None:
            for i, h in enumerate(header):
                if h.upper() == name.upper():
                    return i
            return None

        idx_level   = _col("PROJ LEVEL")
        idx_player  = _col("PLAYER")
        idx_pos     = _col("POS")
        idx_age     = _col("AGE")
        idx_maxlvl  = _col("MAX LEVEL")
        # Column is "Options or R5 Status" in minors sheets
        idx_options = next(
            (i for i, h in enumerate(header) if "OPTIONS" in h.upper()),
            None,
        )

        if idx_player is None:
            continue

        for row in rows[1:]:
            if not row or row[idx_player] is None:
                continue
            player = str(row[idx_player]).strip()
            if not player or player.lower() in ("player", "nan", ""):
                continue

            proj_level = str(row[idx_level]).strip().upper() if idx_level is not None and row[idx_level] else ""
            if proj_level not in levels:
                continue

            pos_raw    = str(row[idx_pos]).strip() if idx_pos is not None and row[idx_pos] else "UNK"
            age        = float(row[idx_age]) if idx_age is not None and row[idx_age] not in (None, "") else np.nan
            max_level  = str(row[idx_maxlvl]).strip() if idx_maxlvl is not None and row[idx_maxlvl] else ""
            opt_val    = row[idx_options] if idx_options is not None else None
            on_40_man  = _is_40man_option(opt_val)

            records.append({
                "Player":      player,
                "pos_raw":     pos_raw,
                "pos_group":   _map_pos(pos_raw),
                "age":         age,
                "proj_level":  proj_level,
                "max_level":   max_level,
                "on_40_man":   on_40_man,
                "depth_sheet": sheet_name,
            })

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Deduplicate across sheets
    df = (
        df.sort_values("proj_level", kind="mergesort")
          .drop_duplicates("Player", keep="first")
          .reset_index(drop=True)
    )
    return df


def merge_depth_with_payroll(
    depth_df: pd.DataFrame,
    roster_status_df: pd.DataFrame,
    league_min_M: float = LEAGUE_MIN_M,
) -> pd.DataFrame:
    """
    Combine depth chart projected players with the team's payroll data.

    For each depth chart player:
      - If found in roster_status_df: use their sal_2026_M / proj_arb_cost_M / war_2025
      - If not found: salary = league_min_M, proj_WAR from depth chart

    Returns DataFrame with columns:
        Player, pos_raw, pos_group, age, proj_WAR, sal_2026_M,
        contract_status, depth_sheet, salary_source
    """
    if depth_df.empty:
        return pd.DataFrame()

    # Build a lookup from player name → payroll row
    payroll_lookup: dict = {}
    if not roster_status_df.empty and "Player" in roster_status_df.columns:
        for _, row in roster_status_df.iterrows():
            payroll_lookup[str(row["Player"]).strip()] = row

    merged_rows = []
    for _, dc_row in depth_df.iterrows():
        player = str(dc_row["Player"]).strip()
        pay_row = payroll_lookup.get(player)

        if pay_row is not None:
            # Use payroll data
            status = str(pay_row.get("contract_status", "locked"))
            if status == "locked":
                sal = float(pay_row["sal_2026_M"]) if not pd.isna(pay_row.get("sal_2026_M")) else league_min_M
            elif status == "arb":
                sal = float(pay_row["proj_arb_cost_M"]) if not pd.isna(pay_row.get("proj_arb_cost_M")) else league_min_M
            else:
                sal = league_min_M

            war = float(pay_row["war_2025"]) if not pd.isna(pay_row.get("war_2025")) else (
                float(dc_row["proj_WAR"]) if not pd.isna(dc_row.get("proj_WAR")) else 0.0
            )
            age = float(pay_row.get("age", dc_row.get("age", np.nan)))
            salary_source = "Payroll"
            pos_group = str(pay_row.get("pos_group", dc_row["pos_group"]))
        else:
            # Not on payroll sheets → league minimum
            status = "lg_min"
            sal = league_min_M
            war = float(dc_row["proj_WAR"]) if not pd.isna(dc_row.get("proj_WAR")) else 0.0
            age = float(dc_row.get("age", np.nan)) if not pd.isna(dc_row.get("age")) else np.nan
            salary_source = "League Min"
            pos_group = dc_row["pos_group"]

        merged_rows.append({
            "Player":           player,
            "pos_raw":          dc_row.get("pos_raw", ""),
            "pos_group":        pos_group,
            "age":              age,
            "proj_WAR":         war,
            "sal_2026_M":       sal,
            "contract_status":  status,
            "depth_sheet":      dc_row.get("depth_sheet", ""),
            "salary_source":    salary_source,
        })

    return pd.DataFrame(merged_rows)
